# -*- coding: utf-8 -*-
"""
种子词库生成（10篇文献语料版：9篇中文设计驱动文献 + CMF英文专著 + 4份政策）
- CMF专著为正常语料，不特殊对待；因其为英文，提取按英文逻辑（英文词频），表格显示为中文翻译
- 输出格式与第二次输出一致：一级维度 | 二级维度 | 三级子维度 | 种子词(中文) | 来源文件 | 原文摘录 | 语料频次
"""
import json, re, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC_CMF = "《The Fundamental Principles of CMF Colour, Material and Finish Design》（Liliana Becerra, 2016）"

# ---------- 读取 CMF 专著文本 ----------
cmf_text = open('文献/CMF_Becerra.txt', encoding='utf-8').read()
cmf_low = cmf_text.lower()

def count_term(*terms):
    return sum(len(re.findall(re.escape(t), cmf_low)) for t in terms)

def body_snippet(term, skip=3, n=120):
    idxs = [m.start() for m in re.finditer(re.escape(term.lower()), cmf_low)]
    if not idxs:
        # 尝试词干
        alt = [m.start() for m in re.finditer(re.escape(term.lower().rstrip('s')), cmf_low)]
        idxs = alt
    if not idxs:
        return ''
    idx = idxs[skip] if len(idxs) > skip else idxs[-1]
    start = max(0, idx - 60)
    end = min(len(cmf_text), idx + len(term) + 60)
    return '...' + cmf_text[start:end].replace('\n', ' ') + '...'

# ---------- CMF 专著贡献词条（英文术语 → 中文，频次=英文词频） ----------
# (中文词, 归属三级子维度, 英文术语, 摘录主术语)
cmf_entries = [
    # --- 色彩 ---
    ('色彩设计', '色彩', ['colour design'], 'colour design'),
    ('配色方案', '色彩', ['colour palette', 'colour palettes'], 'colour palette'),
    ('色彩搭配', '色彩', ['colour combination', 'colour combinations'], 'colour combination'),
    ('色调', '色彩', ['tone', 'tones'], 'tone'),
    ('中性色', '色彩', ['neutral tones'], 'neutral tones'),
    ('色彩趋势', '色彩', ['colour trend', 'colour trends'], 'colour trend'),
    ('色彩心理学', '色彩', ['colour psychology', 'colour affinity'], 'colour affinity'),
    ('色彩测量', '色彩', ['colour measuring', 'colour measurement'], 'colour measuring'),
    ('色彩语言', '色彩', ['visual design language', 'design language'], 'design language'),
    ('数字色彩', '色彩', ['digital colour'], 'digital colour'),
    ('色彩体系', '色彩', ['colour system', 'pantone', 'natural colour system'], 'colour system'),
    # --- 材质 ---
    ('材质', '材质', ['material', 'materials'], 'material'),
    ('材质设计', '材质', ['material design'], 'material design'),
    ('触感', '材质', ['soft touch', 'tactile'], 'soft touch'),
    ('纹理', '材质', ['texture', 'textures', 'grain'], 'texture'),
    ('金属材质', '材质', ['metal', 'metals'], 'metal'),
    ('铝合金', '材质', ['aluminium'], 'aluminium'),
    ('不锈钢', '材质', ['stainless steel'], 'stainless steel'),
    ('碳纤维', '材质', ['carbon fibre'], 'carbon fibre'),
    ('皮革', '材质', ['leather'], 'leather'),
    ('木材', '材质', ['wood', 'wooden', 'natural wood'], 'wood'),
    ('木饰面', '材质', ['wood veneer', 'veneer'], 'wood veneer'),
    ('玻璃', '材质', ['glass'], 'glass'),
    ('塑料', '材质', ['plastic', 'plastics'], 'plastic'),
    ('纺织品', '材质', ['textile', 'textiles', 'fabric', 'woven'], 'textile'),
    ('天然材料', '材质', ['natural material', 'natural materials'], 'natural material'),
    ('合成材料', '材质', ['synthetic', 'synthetics'], 'synthetic'),
    ('橡胶弹性体', '材质', ['rubber', 'elastomer', 'elastomers'], 'rubber'),
    # --- 表面处理 ---
    ('表面处理', '表面处理', ['finish', 'finishes', 'finishing'], 'finish'),
    ('表面装饰', '表面处理', ['surface decoration'], 'surface decoration'),
    ('饰面', '表面处理', ['surface finish', 'surface finishes'], 'surface finish'),
    ('抛光', '表面处理', ['polished', 'polishing'], 'polished'),
    ('哑光', '表面处理', ['matte'], 'matte'),
    ('光泽', '表面处理', ['gloss', 'glossy', 'gloss level'], 'gloss'),
    ('高光', '表面处理', ['high gloss'], 'high gloss'),
    ('缎面', '表面处理', ['satin'], 'satin'),
    ('拉丝', '表面处理', ['brushed', 'brushing'], 'brushed'),
    ('喷砂', '表面处理', ['sandblasted', 'sand-blasting'], 'sandblasted'),
    ('激光加工', '表面处理', ['laser marking', 'laser engraving', 'laser cutting'], 'laser marking'),
    ('涂层', '表面处理', ['coating', 'coatings'], 'coating'),
    ('镀铬', '表面处理', ['chrome'], 'chrome'),
    ('涂料油漆', '表面处理', ['paint', 'paints'], 'paint'),
    ('电化学蚀刻', '表面处理', ['chemical etching', 'etching'], 'chemical etching'),
    ('阳极化处理', '表面处理', ['anodised', 'anodising'], 'anodised'),
    ('模压成型', '表面处理', ['moulding', 'molding'], 'moulding'),
    ('注塑成型', '表面处理', ['injection'], 'injection'),
    ('颜料', '表面处理', ['pigment', 'pigments'], 'pigment'),
    ('染料', '表面处理', ['dye', 'dyes'], 'dye'),
    ('金属闪光效果', '表面处理', ['metallic flake', 'metallic'], 'metallic flake'),
    ('珠光效果', '表面处理', ['pearlescent'], 'pearlescent'),
    ('缝线工艺', '表面处理', ['stitch', 'stitching'], 'stitching'),
    ('编织工艺', '表面处理', ['knit', 'knitting'], 'knit'),
    ('层压工艺', '表面处理', ['laminate', 'laminates'], 'laminates'),
    # --- 美学感知 ---
    ('感知价值', '美学感知', ['perceived value'], 'perceived value'),
    ('奢华感', '美学感知', ['luxury'], 'luxury'),
    ('高端质感', '美学感知', ['premium'], 'premium'),
    ('情感化设计', '美学感知', ['emotional', 'emotion'], 'emotional'),
    ('可持续设计', '美学感知', ['sustainable', 'sustainability'], 'sustainable'),
    ('耐用性', '美学感知', ['durability'], 'durability'),
    ('工艺美感', '美学感知', ['craftsmanship'], 'craftsmanship'),
]

# ---------- 基础 136 词（9 中文文献 + 4 政策） ----------
with open('seed_v2_final.json', encoding='utf-8') as f:
    base = json.load(f)

rows = []

for d1, subs in base.items():
    for d2, subs2 in subs.items():
        for d3, words in subs2.items():
            for w in words:
                src = w['src']
                if '|' in src:
                    typ, name = src.split('|', 1)
                    if typ == '文献':
                        src_name = '《' + name + '》（设计驱动型创新文献）'
                    else:
                        src_name = '《' + name + '》（政策文件）'
                else:
                    src_name = src
                rows.append([d1, d2, d3, w['word'], src_name, w['snippet'], w['freq']])

# ---------- 加入 CMF 词条（正常语料，不特殊标记） ----------
# 跳过与已有136词重复的中文词（含跨维度：同一中文词只保留基础词归属）
existing_words = {r[3] for r in rows}
# 已有词：色彩、新材料、美学、审美、感知价值 等
skip_words = {'色彩', '新材料', '美学', '审美', '感知价值'}
D2 = '色彩、材质与饰面'
for cn, subdim, en_terms, snippet_term in cmf_entries:
    if cn in skip_words:
        continue
    if cn in existing_words:
        continue
    freq = count_term(*en_terms)
    snip = body_snippet(snippet_term)
    rows.append(['形式层（外观）', D2, subdim, cn, SRC_CMF, snip, freq])

# ---------- 按维度结构排序（一级/二级/三级显式顺序）----------
D1_ORDER = ['形式层（外观）', '流程层（设计组织载体维度）', '战略层（设计驱动型创新）']
D2_ORDER = ['造型与结构', '色彩、材质与饰面', '组织载体', '设计人力', '管理与方法',
            '外部协同', '资源投入', '设计战略化', '意义创新', '诠释网络与设计话语']
D3_ORDER = ['外观与造型', '结构与一体化', '人机与交互', '色彩', '材质', '表面处理', '美学感知',
            '设计机构', '平台与载体', '设计人才', '人才培养', '设计管理', '设计方法', '设计工具与数据',
            '协同合作', '设计服务', '经费投入', '研发活动', '创新成果',
            '战略定位', '理念与文化', '能力与升级',
            '产品意义与语言', '情感与体验', '文化符号', '场景与生活方式',
            '创新网络', '需求洞察', '品牌与市场']

def dim_key(r):
    return (D1_ORDER.index(r[0]) if r[0] in D1_ORDER else 99,
            D2_ORDER.index(r[1]) if r[1] in D2_ORDER else 99,
            D3_ORDER.index(r[2]) if r[2] in D3_ORDER else 99,
            r[3])

rows.sort(key=dim_key)

# ---------- 生成 Excel ----------
# 先统计每个三级子维度的去重词数（同一子维度内不重复计）
from collections import Counter
subdim_cnt = Counter()
seen = set()
for r in rows:
    key = (r[0], r[1], r[2], r[3])
    if key in seen:
        continue
    seen.add(key)
    subdim_cnt[(r[0], r[1], r[2])] += 1

def subdim_label(d1, d2, d3):
    n = subdim_cnt.get((d1, d2, d3), 0)
    return f'{d3}({n})' if d3 else d3

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '种子词库（含CMF专著）'

header = ['一级维度', '二级维度', '三级子维度(匹配词数)', '种子词', '来源文件', '原文摘录（词前后各60字）', '语料频次']
ws.append(header)

head_fill = PatternFill('solid', fgColor='4472C4')
head_font = Font(bold=True, color='FFFFFF', size=11)
thin = Side(style='thin', color='B0B0B0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c in range(1, len(header)+1):
    cell = ws.cell(row=1, column=c)
    cell.fill = head_fill
    cell.font = head_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

cur_d1 = cur_d2 = cur_d3 = None
for r in rows:
    d1, d2, d3, w, src, snip, freq = r
    d3_lbl = subdim_label(d1, d2, d3)
    show_d1 = d1 if d1 != cur_d1 else ''
    show_d2 = d2 if (d2 != cur_d2 or d1 != cur_d1) else ''
    show_d3 = d3_lbl if (d3_lbl != cur_d3 or d2 != cur_d2 or d1 != cur_d1) else ''
    ws.append([show_d1, show_d2, show_d3, w, src, snip, freq])
    r_idx = ws.max_row
    for c in range(1, 8):
        cell = ws.cell(row=r_idx, column=c)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=(c in (5, 6)))
    cur_d1, cur_d2, cur_d3 = d1, d2, d3_lbl

widths = [22, 20, 18, 14, 44, 60, 10]
for i, wd in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = wd
ws.freeze_panes = 'A2'

out = '种子词库_第三次输出.xlsx'
wb.save(out)

# 统计
from collections import Counter
d1c = Counter(r[0] for r in rows)
cmf_cnt = sum(1 for r in rows if SRC_CMF in r[4])
print('saved:', out)
print('总词数:', len(rows))
print('CMF专著来源词数:', cmf_cnt)
for k, v in d1c.items():
    print(f'  {k}: {v}')
# 色彩材质饰面维度详情
print()
print('=== 色彩、材质与饰面 维度（含CMF词）===')
for r in rows:
    if r[1] == '色彩、材质与饰面':
        print(f'  {r[2]:8s} {r[3]:10s} 频次:{r[6]}')
