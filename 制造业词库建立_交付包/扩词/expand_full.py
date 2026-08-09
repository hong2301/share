# -*- coding: utf-8 -*-
"""
完整扩词任务：用最优方法（BERT为主 + Word2Vec补充）基于词典.xlsx 47个种子词扩词
- BERT: bert-base-chinese 预训练，语义邻居（主）
- Word2Vec: 语料从头训练，领域共现补充（仅取高质量词）
- 黑名单过滤 + 残词过滤
输出：扩词结果.xlsx（含明细/汇总/合并词库 三个sheet）
"""
import os, json, re, sys
import jieba
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = r'C:\Users\Administrator\Desktop\ck\制造业词库建立'
CORPUS = os.path.join(BASE, '第二次输出', '语料_中文全量.txt')
DICT_XLSX = os.path.join(BASE, '第三次输出', '词典.xlsx')
OUT = os.path.join(BASE, '第三次输出', '扩词结果_词典扩展版.xlsx')

# ---------- 黑名单（客户指定21词 + 工程噪音） ----------
BLACKLIST = {'设计产能','设计压力','设计温度','设计流量','设计参数','设计工况','设计寿命',
             '设计图纸','工程设计','勘察设计','设计院','初步设计','施工图设计','设计资质',
             '顶层设计','制度设计','机制设计','股权设计','薪酬设计','联名担保','联名账户'}

# ---------- 1. 读取种子词及其维度 ----------
import openpyxl as opx
wb0 = opx.load_workbook(DICT_XLSX, data_only=True)
ws0 = wb0.active
seed_info = []  # (词, 层级, 一级维度)
for row in ws0.iter_rows(min_row=2, values_only=True):
    lvl, d1, words = row[0], row[1], row[2]
    if words:
        for w in str(words).replace('，', ',').replace('、', ',').split(','):
            w = w.strip()
            if w and not any(w == s for s, _, _ in seed_info):
                seed_info.append((w, lvl, d1))
seeds = [s for s, _, _ in seed_info]
print('种子词:', len(seeds))

# ---------- 2. 语料 & 分词 ----------
with open(CORPUS, encoding='utf-8') as f:
    corpus = f.read()
for s in seeds:
    jieba.add_word(s)
sentences_raw = [s.strip() for s in re.split(r'[。！？；\n]', corpus) if len(s.strip()) >= 5]
tokenized = [list(jieba.cut(s)) for s in sentences_raw]

from collections import Counter
vocab_counter = Counter()
for toks in tokenized:
    vocab_counter.update(toks)

def clean_word(w):
    """过滤残词/噪音：仅保留2-6字中文词，无叠字、无重复字符"""
    if not re.fullmatch(r'[\u4e00-\u9fff]{2,6}', w):
        return False
    if w in BLACKLIST:
        return False
    # 叠字/残词：如"设设""业是"（字符重复或首尾残缺特征）
    if any(w[i] == w[i+1] for i in range(len(w)-1)):
        return False
    # 无意义词黑名单（小语料常见噪声）
    NOISE = {'二是','三是','四是','其一','其二','其三','本文','文中','其中','此外','同时','以及','一些','因此','由于','对于','进行','通过','如果','那么','这样','那样','可以','应该','必须','比较','非常','进行','以来','以后','以上','以下','其中','目前','如今','日前','如此','为此','一旦','一旦','方面','程度','情况','问题','关系','部分','时期','期间','过程','程中','环节','领域','方面','角度','层面','视角','基础','前提','条件','方式','方法','路径','途径','手段','渠道','概念','范畴','要素','因素','变量','指标','数据','结果','结论','建议','对策','政策','措施','机制','制度','体系','框架','模式','模型','结构','功能','作用','影响','效果','效率','效益','价值','意义','内涵','外延','特征','特点','性质','属性','形态','状态','趋势','方向','目标','目的','任务','内容','形式','种类','类型','类别','系列','方面','上面','下面','上面','之间','之内','之外','上下','左右','前后','等等','以及','或者','还是','可能','能够','需要','要求','希望','实现','达到','取得','获得','产生','形成','促进','推动','带动','提升','提高','增强','加强','加大','强化','优化','完善','改进','改善','发展','建设','构建','建立','创建','设立','成立','形成','增长','增加','扩大','延伸','拓展','深化','丰富','补充','完善','规范','标准','统一','整合','融合','结合','配合','协调','协作','合作','联合','共享','开放','创新','研发','设计','开发','生产','制造','加工','应用','运用','利用','采用','使用','推广','普及','覆盖','涉及','包含','包括','组成','构成','分为','分为','属于','成为','作为','进入','加入','参与','开展','实施','执行','落实','贯彻','推进','部署','安排','组织','管理','监督','检查','审查','考核','评估','评价','认定','审核','批准','同意','支持','鼓励','引导','帮助','协助','服务','供给','需求','消费','购买','销售','市场','客户','用户','顾客','产品','商品','服务','企业','公司','机构','组织','单位','部门','行业','产业','领域','地区','区域','地方','国家','政府','社会','公众','个人','团体','群体','群众','人民','民族','文化','艺术','科学','技术','工程','工业','农业','商业','经济','金融','资本','投资','融资','资金','财产','资产','资源','能源','材料','原料','部件','零件','组件','模块','系统','网络','平台','中心','基地','园区','项目','工程','任务','计划','规划','方案','报告','文件','文档','资料','信息','知识','经验','实践','理论','研究','分析','调查','统计','计算','测量','评估','判断','决策','选择','确定','认定','确认','确保','保障','保护','维护','保持','持续','继续','不断','日益','逐渐','逐步','加快','迅速','快速','有效','显著','明显','突出','重要','关键','核心','重点','优先','首先','其次','最后','最终','初步','基本','全面','系统','整体','综合','总体','共同','普遍','广泛','深入','充分','完全','极其','十分','非常','特别','尤其','更加','更为','越发','愈来','此外','另外','同时','与此','随之','继而','进而','从而','因而','因此','所以','于是','然后','接着','后来','此前','此前','之前','之后','当时','此时','此刻','现在','将来','未来','近期','长期','短期','中期','年度','年份','月份','日期','时间','时刻','时期','阶段','步骤','程序','流程','环节','顺序','进度','过程','历程','经历','体验','感受','感觉','印象','认识','理解','看法','观点','意见','建议','设想','构思','创意','想法','思路','途径','办法','策略','战略','战术','方针','政策','法规','法律','条例','规定','规则','准则','细则','办法','方案','标准','指标','指数','系数','比率','比例','概率','频次','频率','规模','范围','幅度','程度','水平','质量','数量','总量','总额','共计','合计','平均','均值','中位','众数','方差','分布','结构','构成','组成','内容','形式','形态','样式','外观','外形','形状','尺寸','大小','长短','高低','厚薄','轻重','快慢','强弱','多少','有无','是否','与否','能否','可否','以及','之类','等等','其间','之中','之内','之外','以上','以下','如前','如上','下述','上述','前述','上文','下文','全文','本段','本节','本章','本部分','第一部分','第二部分','第三部分','具体来说','总的来看','总体而言','综合来看','综上所述','由此可见','换言之','也就是说','换句话说','换而言之','与此同时','另一方面','除此之外','与此同时','在此基础','在此基础上','相比之下','与此相反','与此不同','与之相比','相对而言','比较而言','就其而言','就其本身','归根结底','从根本','根本来说','本质上','实质上','实际上','事实上','理论上','实践上','总体上说','大体上','基本上','原则上','严格来说','准确来说','确切地说','坦率地说','老实说','说白了','简单说','简要地说','通俗地说','总的来说','概括地说','一言以蔽之'}
    if w in NOISE:
        return False
    return True

candidate_words = [w for w, c in vocab_counter.items() if clean_word(w) and c >= 2]
print('候选词(清洗后, ≥2字, 频次≥2):', len(candidate_words))

# ---------- 3. Word2Vec ----------
from gensim.models import Word2Vec
w2v = Word2Vec(tokenized, vector_size=100, window=5, min_count=1, sg=1, epochs=30, workers=1, seed=42)
print('Word2Vec 词表:', len(w2v.wv))

# ---------- 4. BERT ----------
print('加载 BERT (bert-base-chinese)...')
from transformers import AutoTokenizer, AutoModel
import torch
tok = AutoTokenizer.from_pretrained('bert-base-chinese')
model = AutoModel.from_pretrained('bert-base-chinese')
model.eval()

def get_bert_embeddings(sentences_batch, max_len=64):
    enc = tok(sentences_batch, padding=True, truncation=True, max_length=max_len, return_tensors='pt')
    with torch.no_grad():
        out = model(**enc)
    return enc, out.last_hidden_state

word_embs, word_cnt = {}, {}
BATCH = 32
for i in range(0, len(sentences_raw), BATCH):
    batch = sentences_raw[i:i+BATCH]
    enc, hidden = get_bert_embeddings(batch)
    chars_all = enc['input_ids']
    for b in range(len(batch)):
        chars = tok.convert_ids_to_tokens(chars_all[b])
        h = hidden[b]
        toks = list(jieba.cut(batch[b]))
        pos = 1
        for tk in toks:
            if pos >= len(chars) - 1:
                break
            if not re.fullmatch(r'[\u4e00-\u9fff]+', tk):
                pos += max(len(tk), 1)
                continue
            L = len(tk)
            if pos + L > len(chars):
                break
            if all(chars[pos+j] == tk[j] for j in range(L)):
                emb = h[pos:pos+L].mean(dim=0).numpy()
                word_embs.setdefault(tk, []).append(emb)
                word_cnt[tk] = word_cnt.get(tk, 0) + 1
            pos += L
    if (i // BATCH) % 30 == 0:
        print(f'  BERT 进度: {i}/{len(sentences_raw)} 词数: {len(word_embs)}')

bert_vec = {}
for w, embs in word_embs.items():
    if clean_word(w) and word_cnt[w] >= 2:
        bert_vec[w] = np.mean(np.stack(embs), axis=0)
print('BERT 词向量数:', len(bert_vec))

def cos(a, b):
    a = a / (np.linalg.norm(a) + 1e-9)
    b = b / (np.linalg.norm(b) + 1e-9)
    return float(a @ b)

# ---------- 5. 扩词（BERT 主 + W2V 补） ----------
# 预计算所有候选词的 BERT 向量矩阵，供快速检索
def encode_text(text):
    enc = tok([text], padding=True, truncation=True, max_length=32, return_tensors='pt')
    with torch.no_grad():
        out = model(**enc)
    return out.last_hidden_state

def seed_len(w):
    return len(tok.tokenize(w))

# 种子词向量：优先用语料中出现位置聚合；未出现则单独编码（预训练语义）
seed_bert = {}
seed_in_corpus = {}
for seed, _, _ in seed_info:
    if seed in bert_vec:
        seed_bert[seed] = bert_vec[seed]
        seed_in_corpus[seed] = True
    else:
        h = encode_text(seed)
        seed_bert[seed] = h[0, 1:seed_len(seed)+1].mean(dim=0).numpy()
        seed_in_corpus[seed] = False

def bert_expand(seed, topk=20):
    if seed not in seed_bert:
        return []
    sv = seed_bert[seed]
    sims = [(w, cos(sv, v)) for w, v in bert_vec.items() if w != seed]
    # 语料外种子词（单独编码）提高阈值，减少泛词噪音
    thr = 0.60 if seed_in_corpus.get(seed, True) else 0.75
    sims = [(w, s) for w, s in sims if s >= thr]
    sims.sort(key=lambda x: -x[1])
    return sims[:topk]

def w2v_expand(seed, topk=20):
    if seed not in w2v.wv:
        return []
    return [(w, float(s)) for w, s in w2v.wv.most_similar(seed, topn=topk)]

results = []
for seed, lvl, d1 in seed_info:
    be = bert_expand(seed)
    we = w2v_expand(seed)
    # 扩词只用 BERT（最优方法）；Word2Vec 结果单独保留作参考
    bert_items = [(w, s, 'BERT') for w, s in be if clean_word(w)]
    w2v_ref = [(w, s) for w, s in we if clean_word(w)]
    results.append({'seed': seed, 'lvl': lvl, 'd1': d1, 'items': bert_items,
                    'w2v_ref': w2v_ref, 'in_corpus': seed_in_corpus.get(seed, True)})

# ---------- 6. 输出 xlsx ----------
wb = openpyxl.Workbook()

thin = Side(style='thin', color='B0B0B0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill('solid', fgColor='4472C4')
head_font = Font(bold=True, color='FFFFFF', size=11)
seed_fill = PatternFill('solid', fgColor='FFF2CC')
w2v_fill = PatternFill('solid', fgColor='E2EFDA')

# ---- Sheet1: 扩词明细（仅 BERT 扩词） ----
ws1 = wb.active
ws1.title = '扩词明细'
h1 = ['种子词', '层级', '一级维度', '种子词语料状态', '扩词(BERT)', '相似度', '是否保留(删/留)']
ws1.append(h1)
for c in range(1, len(h1)+1):
    cell = ws1.cell(row=1, column=c)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
for r in results:
    if not r['items']:
        ws1.append([r['seed'], r['lvl'], r['d1'], '语料内' if r['in_corpus'] else '语料外', '（未扩出有效词）', '', ''])
        ws1.cell(row=ws1.max_row, column=5).font = Font(color='FF0000', italic=True)
        continue
    for i, (w, s, method) in enumerate(r['items']):
        row = [r['seed'], r['lvl'], r['d1'], '语料内' if r['in_corpus'] else '语料外', w, round(s, 4), '']
        if i > 0:
            row[0] = row[1] = row[2] = row[3] = ''
        ws1.append(row)
        r_idx = ws1.max_row
        for c in range(1, 8):
            cell = ws1.cell(row=r_idx, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical='top')
        if i == 0:
            for c in range(1, 5):
                ws1.cell(row=r_idx, column=c).fill = seed_fill

# ---- Sheet2: 按种子词汇总（BERT扩词 + W2V参考） ----
ws2 = wb.create_sheet('按种子词汇总')
h2 = ['种子词', '层级', '一级维度', '语料状态', 'BERT扩词数', 'BERT扩词(Top10)', 'Word2Vec参考词(不建议直接采用)', '审核意见']
ws2.append(h2)
for c in range(1, len(h2)+1):
    cell = ws2.cell(row=1, column=c)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
for r in results:
    bert_words = [w for w, s, m in r['items'] if m == 'BERT'][:10]
    w2v_words = [w for w, s in r['w2v_ref'][:8]]
    ws2.append([r['seed'], r['lvl'], r['d1'], '语料内' if r['in_corpus'] else '语料外', len(r['items']),
                '、'.join(bert_words), '、'.join(w2v_words), ''])
    r_idx = ws2.max_row
    for c in range(1, 9):
        ws2.cell(row=r_idx, column=c).border = border
        ws2.cell(row=r_idx, column=c).alignment = Alignment(vertical='top', wrap_text=True)

# ---- Sheet3: 合并词库（种子+扩词去重） ----
ws3 = wb.create_sheet('合并词库')
merged = {}
for r in results:
    for w, s, m in r['items']:
        if w not in merged:
            merged[w] = {'seed': r['seed'], 'd1': r['d1'], 'method': m, 'best_sim': s,
                         'from_seeds': [r['seed']]}
        else:
            merged[w]['from_seeds'].append(r['seed'])
            merged[w]['best_sim'] = max(merged[w]['best_sim'], s)
h3 = ['序号', '词', '是否种子词', '来源种子词', '一级维度', '方法', '最佳相似度']
ws3.append(h3)
for c in range(1, len(h3)+1):
    cell = ws3.cell(row=1, column=c)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
idx = 1
for w, info in sorted(merged.items(), key=lambda x: -x[1]['best_sim']):
    is_seed = '是' if w in seeds else '否'
    ws3.append([idx, w, is_seed, '、'.join(dict.fromkeys(info['from_seeds']))[:100],
                info['d1'], info['method'], round(info['best_sim'], 4)])
    r_idx = ws3.max_row
    for c in range(1, 8):
        ws3.cell(row=r_idx, column=c).border = border
    idx += 1

# 列宽
for ws, widths in [(ws1, [14, 10, 22, 12, 18, 10, 18]), (ws2, [16, 10, 22, 10, 8, 60, 40, 14]), (ws3, [6, 16, 10, 60, 22, 8, 12])]:
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = 'A2'

wb.save(OUT)
print('saved:', OUT)

# 统计
n_seed_ok = sum(1 for r in results if r['items'])
n_total = sum(len(r['items']) for r in results)
print(f'可扩词种子词: {n_seed_ok}/{len(results)}')
print(f'总扩词数(去重前): {n_total}')
print(f'合并词库总词数: {len(merged)}')
