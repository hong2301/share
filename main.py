#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
企业查询平台 · 抓取工具（CLI / 界面双模式）
============================================
目标网站：https://skypt.gdcic.net/openplatform/#/web/enterprise
接口已通过 JS 逆向破解（详见 需求整理.md 第七章），全部无需 token/验证码。

用法：
  界面模式（默认）：python main.py
  CLI 模式：        python main.py 关键词1 关键词2 ...

CLI 模式采集后生成 export_时间戳/ 目录，内含合并 xlsx（自动去重）和 分文件/ 子目录。
界面模式（PyQt5）：添加关键词 → 采集 → 预览 → 批量导出。

依赖：requests, openpyxl（界面模式另需 PyQt5）
"""
import os
import re
import sys
import time
import json
import hashlib
import threading
from datetime import datetime
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed

# 检查依赖
try:
    import requests
except ImportError:
    print("错误: 缺少 requests 库")
    print("请执行: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 缺少 openpyxl 库")
    print("请执行: pip install openpyxl")
    sys.exit(1)

# ==================== PyQt5（可选导入，CLI 不依赖） ====================
try:
    from PyQt5.QtWidgets import *
    from PyQt5.QtCore import Qt, pyqtSignal, QObject
    from PyQt5.QtGui import QFont, QColor
    HAS_GUI = True
except ImportError:
    HAS_GUI = False

# ======================= 配置 =======================
BASE_URL = "https://skypt.gdcic.net/api"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Referer": "https://skypt.gdcic.net/openplatform/",
    "Accept": "application/json, text/plain, */*",
}
PAGE_SIZE = 300          # 单次最多返回 300 条（后端限制，前端也只显示 30 页）
MAX_PAGES = 1            # 每关键词每城市抓 1 页 = 300 条
REQUEST_INTERVAL = 0.3   # 请求间隔（秒），避免触发风控
THREADS = 6              # 详情并发线程数
OUTPUT_DIR = "输出"
GUANGDONG_PID = "440000"  # 广东省行政区划代码（城市接口 pid 参数）

STATUS_L = {'pending': '待处理', 'collecting': '采集中', 'done': '采集完毕', 'failed': '采集失败'}
STATUS_C = {'pending': '#8899a6', 'collecting': '#1d9bf0', 'done': '#00ba7c', 'failed': '#f4212e'}

# 城市列表缓存
def get_cities(api):
    """获取广东省 21 个城市列表 [{id, extName}]（id 统一转字符串）"""
    data = api._get("/system/districts/list3", {"pid": GUANGDONG_PID, "pageSize": 999})
    cities = data.get("rows") or []
    for c in cities:
        c["id"] = str(c["id"])
    return cities

# ======================= Excel 表头（自动按字段填充） =======================
# 列顺序：与页面导出示例一致（基本信息在前，证书在后，多余数据加最右边）
COLUMNS = [
    # ── 基本信息 ──
    "企业名称", "所在市", "统一社会信用代码", "注册资本（万元）", "注册类型",
    "法定代表人", "成立时间",
    # ── 安全生产许可证 ──
    "安许证-证书编号", "安许证-发证机关", "安许证-有效期", "安许证-证书状态",
    # ── 企业资质（备案）证书 ──
    "资质-证书编号", "资质-证书类别", "资质-等级", "资质-发证机关",
    "资质-发证日期", "资质-有效期", "资质-证书状态",
    # ── 来源 ──
    "来源关键词",
    # ── 多余数据（加在最右边） ──
    "企业编码", "企业类型", "城市编码",
    "安许证-发证日期", "安许证-经营范围",
    "资质-证书类别编码", "资质-经营范围",
]

# 列表接口 → 需求字段 映射
LIST_FIELD_MAP = {
    "entName": "企业名称",
    "city": "所在市",
    "creditCode": "统一社会信用代码",
    "regCapital": "注册资本（万元）",
    "regType": "注册类型",
    "legalName": "法定代表人",
    "foundDate": "成立时间",
    "entCode": "企业编码",
    "entType": "企业类型",
    "cityCode": "城市编码",
}

# 安许证 / 资质证书 → 需求字段 映射
SAFETY_FIELD_MAP = {
    "certNum": "安许证-证书编号",
    "issueOrg": "安许证-发证机关",
    "issueDate": "安许证-发证日期",
    "expireDate": "安许证-有效期",
    "certStatus": "安许证-证书状态",
    "scope": "安许证-经营范围",
}
QUA_FIELD_MAP = {
    "certNum": "资质-证书编号",
    "certType": "资质-证书类别",
    "certTypeCode": "资质-证书类别编码",
    "quaLevel": "资质-等级",
    "issueOrg": "资质-发证机关",
    "issueDate": "资质-发证日期",
    "expireDate": "资质-有效期",
    "certStatus": "资质-证书状态",
    "scope": "资质-经营范围",
}

# ======================= API 封装 =======================
class ApiClient:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)

    def _get(self, path, params=None, retries=3):
        """带重试的 GET 请求"""
        for attempt in range(retries):
            try:
                r = self.session.get(f"{BASE_URL}{path}", params=params, timeout=20)
                data = r.json()
                if data.get("code") == 0:
                    return data
                raise RuntimeError(f"接口返回错误: {data.get('msg')}")
            except Exception as e:
                if attempt == retries - 1:
                    raise
                time.sleep(2 * (attempt + 1))
        return None

    def get_enterprise_list(self, ent_name, page_num=1, page_size=PAGE_SIZE, city_code=None):
        """企业列表：entName 关键词搜索（可选 cityCode 按城市细分，突破 300 条限制）"""
        params = {
            "entName": ent_name, "pageNum": page_num, "pageSize": page_size,
        }
        if city_code:
            params["cityCode"] = city_code
        return self._get("/openplatform/enterpriseInGd/list", params)

    def get_cert_safety(self, ent_code):
        """安全生产许可证"""
        return self._get("/openplatform/enterpriseCertSafety/list", {"entCode": ent_code})

    def get_cert_qua(self, ent_code, page_num=1, page_size=10):
        """企业资质（备案）证书"""
        return self._get("/openplatform/enterpriseCertQua/list", {
            "entCode": ent_code, "pageNum": page_num, "pageSize": page_size,
        })


# ======================= 数据处理 =======================
def join_values(rows, field_map, max_items=5):
    """把多条证书记录的同名字段用分号拼接成单列值"""
    if not rows:
        return {v: "" for v in field_map.values()}
    result = {}
    for cn_field, col_name in field_map.items():
        values = [str(r.get(cn_field) or "") for r in rows[:max_items]]
        values = [v for v in values if v]  # 去掉空值
        result[col_name] = "；".join(values)
    return result


def build_row(ent, safety_rows, qua_rows, keyword):
    """构建一行 Excel 数据"""
    row = {}
    # 基本信息
    for api_field, col_name in LIST_FIELD_MAP.items():
        row[col_name] = ent.get(api_field) or ""
    # 安许证（多条拼接）
    row.update(join_values(safety_rows, SAFETY_FIELD_MAP))
    # 资质证书（多条拼接）
    row.update(join_values(qua_rows, QUA_FIELD_MAP))
    # 来源关键词
    row["来源关键词"] = keyword
    return row


def row_fingerprint(row):
    """全字段指纹：所有字段都重复 → 认定为重复"""
    s = "|".join(f"{k}={v}" for k, v in row.items())
    return hashlib.md5(s.encode("utf-8")).hexdigest()


# ======================= 抓取逻辑 =======================
def fetch_ent_detail(api, ent):
    """获取单个企业的详情（安许证 + 资质证书），供并发调用"""
    ent_code = ent.get("entCode") or ""
    try:
        safety = (api.get_cert_safety(ent_code).get("rows") or []) if ent_code else []
        time.sleep(REQUEST_INTERVAL)
        qua = (api.get_cert_qua(ent_code).get("rows") or []) if ent_code else []
        time.sleep(REQUEST_INTERVAL)
    except Exception as e:
        print(f"    ⚠ 详情获取失败 {ent.get('entName')}: {e}")
        safety, qua = [], []
    return build_row(ent, safety, qua, "")


def fetch_keyword(api, keyword, progress_cb=None):
    """抓取单个关键词的所有数据

    策略：
      - 关键词命中 ≤300 条 → 一次拿完
      - 命中 >300 条 → 按 21 个城市细分抓取（每城市独立 300 条上限），
        汇总后按 entCode 去重，突破单关键词 300 条限制
    """
    # 第一步：不带城市查询，看总量
    data = api.get_enterprise_list(keyword, page_num=1, page_size=PAGE_SIZE)
    total = data.get("total", 0)
    ents = data.get("rows") or []
    if progress_cb:
        progress_cb(keyword, "全广东", total, len(ents))

    if total <= PAGE_SIZE:
        # 一次拿满，直接进详情
        return _fetch_details(api, ents, keyword, progress_cb)

    # 第二步：按城市细分
    cities = get_cities(api)
    seen_codes, all_ents = set(), []
    for city in cities:
        try:
            d = api.get_enterprise_list(keyword, page_num=1, page_size=PAGE_SIZE, city_code=city["id"])
            rows = d.get("rows") or []
        except Exception as e:
            print(f"    ⚠ 城市 {city['extName']} 查询失败: {e}")
            continue
        if progress_cb:
            progress_cb(keyword, city["extName"], d.get("total", 0), len(rows))
        for ent in rows:
            code = ent.get("entCode")
            if code and code in seen_codes:
                continue
            if code:
                seen_codes.add(code)
            all_ents.append(ent)
        time.sleep(REQUEST_INTERVAL)
    return _fetch_details(api, all_ents, keyword, progress_cb)


def _fetch_details(api, ents, keyword, progress_cb=None):
    """并发获取一批企业的详情（安许证 + 资质证书）"""
    rows = []
    detail_rows = [None] * len(ents)
    with ThreadPoolExecutor(max_workers=THREADS) as pool:
        futures = {pool.submit(fetch_ent_detail, api, ent): i for i, ent in enumerate(ents)}
        for fut in as_completed(futures):
            idx = futures[fut]
            detail_rows[idx] = fut.result()
    for row in detail_rows:
        if row:
            row["来源关键词"] = keyword
            rows.append(row)
    return rows


# ======================= Excel 导出 =======================
def write_xlsx_sheet(ws, rows):
    """按 COLUMNS 填充一个 sheet（GUI/CLI 共用），表头自动填充"""
    if not rows:
        return
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    halign = Alignment(horizontal="center", vertical="center")

    for c, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = halign

    for r, row in enumerate(rows, 2):
        for c, col in enumerate(COLUMNS, 1):
            # 接口返回什么存什么（页面为准，不做字典翻译）
            ws.cell(row=r, column=c, value=row.get(col, ""))

    # 列宽自适应（中文按 2 字符宽估算）
    for c, col in enumerate(COLUMNS, 1):
        max_len = len(col) * 2
        for r in range(2, min(len(rows) + 2, 200)):  # 采样前 200 行
            v = ws.cell(row=r, column=c).value
            if v:
                v = str(v)
                max_len = max(max_len, len(v) * 2 if any('\u4e00' <= ch <= '\u9fff' for ch in v) else len(v))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"  # 冻结表头


def write_xlsx_file(path, rows, sheet_title="企业数据"):
    """写单个 xlsx 文件，失败时抛出异常"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        write_xlsx_sheet(ws, rows)
        dir_path = os.path.dirname(path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        wb.save(path)
        return path
    except ImportError:
        raise RuntimeError("缺少 openpyxl 库，请执行: pip install openpyxl")
    except PermissionError:
        raise RuntimeError(f"文件被占用或无权限写入: {path}\n请关闭已打开的 Excel 文件后重试")
    except Exception as e:
        raise RuntimeError(f"写入 Excel 失败: {e}")


def dedup_rows(rows_by_kw, keywords):
    """汇总 + 全字段去重，返回 (all_rows, dup_count)"""
    all_rows, seen = [], set()
    dup = 0
    for kw in keywords:
        for row in rows_by_kw.get(kw, []):
            fp = row_fingerprint(row)
            if fp not in seen:
                seen.add(fp)
                all_rows.append(row)
            else:
                dup += 1
    return all_rows, dup


# ======================= 节点缓存（断点续传） =======================
# 采集过程可能很长（单关键词最久十几分钟），每个关键词完成后立即落盘；
# 中断/失败后重跑同一命令，已缓存的关键词自动跳过（--fresh 强制全量重采）。
CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "采集缓存")


def _cache_path(kw):
    safe = re.sub(r'[\\/:*?"<>|\r\n]', '_', kw)[:50]
    return os.path.join(CACHE_DIR, f"{safe}.json")


def load_cache(kw):
    """读取某关键词的已采集结果，无缓存返回 None"""
    p = _cache_path(kw)
    if not os.path.exists(p):
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def save_cache(kw, rows):
    """节点保存：关键词采集结果立即落盘（JSON）"""
    os.makedirs(CACHE_DIR, exist_ok=True)
    with open(_cache_path(kw), "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False)


def clear_cache(keywords):
    """删除指定关键词的缓存（--fresh 时用）"""
    for kw in keywords:
        p = _cache_path(kw)
        if os.path.exists(p):
            os.remove(p)


# ======================= 核心查询入口（GUI / CLI 共用） =======================
def do_search(keyword, progress_cb=None):
    """抓取单个关键词，返回 rows 列表"""
    api = ApiClient()
    return fetch_keyword(api, keyword, progress_cb)


# ======================= CLI 模式 =======================
def run_cli(keywords, fresh=False):
    """CLI 模式：传入关键词列表，采集后输出合并 + 分文件 xlsx

    fresh=False 时断点续传：已节点缓存的关键词直接跳过（显示条数），
    fresh=True 时强制全量重采（先清缓存）。
    """
    print("企业查询平台 (CLI)")
    print("关键词: " + ", ".join(keywords))
    print()

    if fresh:
        clear_cache(keywords)
        print("(--fresh 已清空缓存，全量重采)")
        print()

    results = {}
    total = len(keywords)
    for i, kw in enumerate(keywords):
        if not fresh:
            cached = load_cache(kw)
            if cached is not None:
                print(f"[{i+1}/{total}] {kw}: 已缓存 {len(cached)} 条，跳过（--fresh 可强制重采）")
                results[kw] = cached
                continue

        print(f"[{i+1}/{total}] 查询: {kw}")

        def cb(k, scope, t, got):
            print(f"    [{scope}] 共 {t} 条，本次返回 {got} 条")

        try:
            data = do_search(kw, cb)
            save_cache(kw, data)  # 节点保存：完成即落盘，中断可续传
            results[kw] = data
            print(f"  -> {len(data)} 条（已保存节点）")
        except Exception as e:
            print(f"  -> 失败: {e}")
            results[kw] = []

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    # CLI 模式输出到 exe/脚本所在目录
    base_dir = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
    base = os.path.join(base_dir, f"export_{ts}")
    os.makedirs(base, exist_ok=True)

    # 合并文件（自动去重）
    all_rows, dup = dedup_rows(results, keywords)
    if all_rows:
        merge_path = os.path.join(base, f"合并_{ts}.xlsx")
        write_xlsx_file(merge_path, all_rows)
        raw = sum(len(results.get(k, [])) for k in keywords)
        print(f"合并文件: {merge_path}")
        print(f"   原始 {raw} 条，去重后 {len(all_rows)} 条（去除 {dup} 条重复）")

    # 分文件
    sub_dir = os.path.join(base, "分文件")
    os.makedirs(sub_dir, exist_ok=True)
    for kw in keywords:
        data = results.get(kw, [])
        if data:
            safe_name = re.sub(r'[\\/:*?"<>|\r\n]', '_', kw)[:50]
            fpath = os.path.join(sub_dir, f"{safe_name}_{ts}.xlsx")
            write_xlsx_file(fpath, data)
    print(f"分文件目录: {sub_dir}")
    print(f"完成! 共 {len(all_rows)} 条记录（去重后）")


# ======================= GUI 信号 =======================
class Sigs(QObject):
    status = pyqtSignal(str, str)
    progress = pyqtSignal(str, int)   # 关键词, 采集中间数量
    done = pyqtSignal(str, list)
    fail = pyqtSignal(str, str)
    all_done = pyqtSignal()


if HAS_GUI:
    sigs = Sigs()
else:
    sigs = None


# ======================= GUI 主窗口 =======================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("企业查询平台")
        self.setGeometry(100, 100, 1500, 820)
        self.setMinimumSize(1100, 600)

        self.keywords = []           # 关键词列表（有序）
        self.tasks = {}              # {关键词: {status, count, data, error}}
        self.collecting = False      # 采集进行中
        self._cancel = False         # 取消采集标志
        self.current_view = None     # 当前预览的关键词

        # 连接信号
        sigs.status.connect(self._on_status)
        sigs.progress.connect(self._on_progress)
        sigs.done.connect(self._on_done)
        sigs.fail.connect(self._on_fail)
        sigs.all_done.connect(self._on_all_done)

        self._build()
        self._update_buttons()

    # ── 按钮工厂 ──
    def _mk_btn(self, text, bg, fg='#fff', border=None, small=False):
        b = QPushButton(text)
        if small:
            color = fg if fg and fg != '#fff' else bg
            b.setStyleSheet(
                f"QPushButton{{background:transparent;color:{color};border:none;"
                f"padding:2px 6px;font-size:12px;}}"
                f"QPushButton:hover{{text-decoration:underline}}"
            )
            b.setCursor(Qt.PointingHandCursor)
        else:
            b.setStyleSheet(
                f"QPushButton{{background:{bg};color:{fg};border:none;"
                f"padding:8px 18px;border-radius:4px;font-weight:600;font-size:14px;min-width:70px}}"
                f"QPushButton:hover{{opacity:0.85}}"
                f"QPushButton:disabled{{background:#ccc;color:#888}}"
            )
            b.setFixedHeight(40)
        return b

    # ── 界面构建 ──
    def _build(self):
        cw = QWidget()
        self.setCentralWidget(cw)
        root = QVBoxLayout(cw)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # 顶栏
        bar = QHBoxLayout()
        bar.setSpacing(8)
        bar.addWidget(QLabel("关键词"))
        self.kw_input = QLineEdit()
        self.kw_input.setPlaceholderText("输入后回车添加，支持空格/逗号/顿号/分号分隔")
        self.kw_input.setFixedWidth(380)
        self.kw_input.setFixedHeight(38)
        self.kw_input.returnPressed.connect(self.add_keyword)
        bar.addWidget(self.kw_input)

        self.btn_add = self._mk_btn("添加", '#fff', '#1d9bf0', '#1d9bf0')
        self.btn_add.clicked.connect(self.add_keyword)
        bar.addWidget(self.btn_add)

        self.btn_import = self._mk_btn("导入", '#fff', '#536471', '#8899a6')
        self.btn_import.clicked.connect(self.import_keywords)
        bar.addWidget(self.btn_import)

        bar.addSpacing(12)

        self.btn_start = self._mk_btn("▶ 采集", '#1d9bf0')
        self.btn_start.clicked.connect(self.toggle_collect)
        bar.addWidget(self.btn_start)

        self.btn_export = self._mk_btn("批量导出", '#00ba7c')
        self.btn_export.clicked.connect(self.export_all)
        self.btn_export.setEnabled(False)
        bar.addWidget(self.btn_export)

        bar.addStretch()

        self.btn_clear = self._mk_btn("清空", '#fff', '#f4212e', '#f4212e')
        self.btn_clear.clicked.connect(self.clear_all)
        bar.addWidget(self.btn_clear)

        root.addLayout(bar)

        # 分割器
        splitter = QSplitter(Qt.Vertical)
        root.addWidget(splitter, 1)

        # 关键词表格
        self.task_table = QTableWidget(0, 4)
        self.task_table.setHorizontalHeaderLabels(['关键词', '状态', '数量', '操作'])
        self.task_table.setColumnWidth(0, 180)
        self.task_table.setColumnWidth(1, 120)
        self.task_table.setColumnWidth(2, 80)
        self.task_table.setColumnWidth(3, 220)
        self.task_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.task_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.task_table.setSelectionMode(QTableWidget.NoSelection)
        self.task_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.task_table.verticalHeader().setVisible(False)
        self.task_table.setShowGrid(True)
        self.task_table.setAlternatingRowColors(True)
        for col in range(4):
            item = self.task_table.horizontalHeaderItem(col)
            if item:
                item.setTextAlignment(Qt.AlignCenter)
        tf = QFont("Microsoft YaHei"); tf.setPixelSize(13)
        self.task_table.setFont(tf)
        self.task_table.horizontalHeader().setFont(tf)
        splitter.addWidget(self.task_table)

        # 数据预览
        self.preview_box = QGroupBox("数据预览")
        pv = QVBoxLayout(self.preview_box)
        pbar = QHBoxLayout()
        self.preview_label = QLabel("")
        pbar.addWidget(self.preview_label)
        pbar.addStretch()
        btn_xlsx = self._mk_btn("导出 XLSX", '#00ba7c', small=True)
        btn_xlsx.clicked.connect(self.export_single)
        pbar.addWidget(btn_xlsx)
        pv.addLayout(pbar)

        self.data_table = QTableWidget(0, 1)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.data_table.verticalHeader().setVisible(False)
        self.data_table.setShowGrid(True)
        self.data_table.horizontalHeader().setStretchLastSection(False)
        pv.addWidget(self.data_table)

        self.preview_box.setVisible(False)
        tf2 = QFont("Microsoft YaHei"); tf2.setPixelSize(12)
        self.data_table.setFont(tf2)
        self.data_table.horizontalHeader().setFont(tf2)
        splitter.addWidget(self.preview_box)
        splitter.setSizes([360, 300])

    # ── 关键词管理 ──
    def add_keyword(self):
        raw = self.kw_input.text().strip()
        if not raw:
            return
        parts = re.split(r'[\s,，、;；]+', raw)
        parts = list(dict.fromkeys(p.strip() for p in parts if p.strip()))  # 去重保序
        if not parts:
            QMessageBox.warning(self, "提示", "输入错误")
            return
        added = 0
        for kw in parts:
            if kw in self.keywords:
                continue
            self.keywords.append(kw)
            self.tasks[kw] = {'status': 'pending', 'count': 0, 'data': None, 'error': None}
            added += 1
        self.kw_input.clear()
        if added == 0 and parts:
            QMessageBox.warning(self, "提示", "关键词已存在")
        else:
            self._render()
            self._update_buttons()

    def import_keywords(self):
        """批量导入关键词：选择文件或粘贴文本"""
        if self.collecting:
            QMessageBox.warning(self, "提示", "采集中，无法导入")
            return
        box = QMessageBox(self)
        box.setWindowTitle("导入方式")
        box.setText("选择导入方式：")
        file_btn = box.addButton("从文件导入", QMessageBox.YesRole)
        text_btn = box.addButton("粘贴文本导入", QMessageBox.NoRole)
        cancel_btn = box.addButton("取消", QMessageBox.RejectRole)
        box.exec_()
        clicked = box.clickedButton()
        if clicked == cancel_btn:
            return

        if clicked == file_btn:
            path, _ = QFileDialog.getOpenFileName(
                self, "选择关键词文件", "",
                "文本文件 (*.txt *.csv *.tsv);;所有文件 (*)"
            )
            if not path:
                return
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    text = f.read()
            except Exception:
                try:
                    with open(path, 'r', encoding='gbk') as f:
                        text = f.read()
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"读取文件失败: {e}")
                    return
        else:
            # 粘贴文本
            text, ok = QInputDialog.getMultiLineText(
                self, "粘贴关键词", "每行一个，或用空格/逗号/顿号分隔："
            )
            if not ok or not text.strip():
                return

        parts = re.split(r'[\s,，、;；\n]+', text.strip())
        parts = list(dict.fromkeys(p.strip() for p in parts if p.strip()))
        if not parts:
            QMessageBox.warning(self, "提示", "未找到有效关键词")
            return
        added = 0
        for kw in parts:
            if kw in self.keywords:
                continue
            self.keywords.append(kw)
            self.tasks[kw] = {'status': 'pending', 'count': 0, 'data': None, 'error': None}
            added += 1
        if added == 0:
            QMessageBox.warning(self, "提示", "所有关键词已存在")
        else:
            self._render()
            self._update_buttons()
            QMessageBox.information(self, "提示", f"导入 {added} 个关键词")

    def clear_all(self):
        if self.collecting:
            QMessageBox.warning(self, "提示", "采集中，无法清空")
            return
        box = QMessageBox(self)
        box.setWindowTitle("确认")
        box.setText("清空所有关键词？")
        yes_btn = box.addButton("是", QMessageBox.YesRole)
        no_btn = box.addButton("否", QMessageBox.NoRole)
        box.exec_()
        if box.clickedButton() != yes_btn:
            return
        self.keywords.clear()
        self.tasks.clear()
        self.current_view = None
        self._hide_preview()
        self._render()
        self._update_buttons()

    def remove_kw(self, kw):
        if self.collecting:
            return
        self.keywords.remove(kw)
        self.tasks.pop(kw, None)
        if self.current_view == kw:
            self.current_view = None
            self._hide_preview()
        self._render()
        self._update_buttons()

    # ── 采集 ──
    def toggle_collect(self):
        if self.collecting:
            # 采集中点击 → 取消
            self._cancel = True
            self.collecting = False
            self._update_buttons()
            return
        pending = [k for k in self.keywords
                   if self.tasks.get(k, {}).get('status') in ('pending', 'failed')]
        if not pending:
            QMessageBox.information(self, "提示", "没有待处理或失败的关键词")
            return
        self.collecting = True
        self._cancel = False
        self._update_buttons()
        for k in pending:
            self.tasks[k]['status'] = 'pending'
            self.tasks[k]['error'] = None
        self._render()
        t = threading.Thread(target=self._run, args=(pending,), daemon=True)
        t.start()

    def _run(self, keywords):
        for kw in keywords:
            if self._cancel:
                break
            sigs.status.emit(kw, 'collecting')
            print(f"🔍 开始采集: {kw}")
            try:
                count = [0]
                def cb(k, scope, t, got):
                    count[0] += got
                    sigs.progress.emit(k, count[0])
                data = do_search(kw, progress_cb=cb)
                save_cache(kw, data)
                sigs.done.emit(kw, data)
                print(f"  ✅ {kw}: {len(data)} 条（节点已保存）")
            except Exception as e:
                sigs.fail.emit(kw, str(e))
                print(f"  ❌ {kw}: 失败 {e}")
        sigs.all_done.emit()

    def _on_status(self, kw, st):
        if kw in self.tasks:
            self.tasks[kw]['status'] = st
        self._render()

    def _on_progress(self, kw, count):
        """采集中实时更新数量"""
        if kw in self.tasks:
            self.tasks[kw]['count'] = count
            self._render()

    def _on_done(self, kw, data):
        self.tasks[kw] = {'status': 'done', 'count': len(data), 'data': data, 'error': None}
        self._render()
        if self.current_view == kw:
            self._show_data(kw)

    def _on_fail(self, kw, err):
        self.tasks[kw]['status'] = 'failed'
        self.tasks[kw]['error'] = err
        self._render()

    def _on_all_done(self):
        self.collecting = False
        self._cancel = False
        self._update_buttons()

    def retry_one(self, kw):
        if self.collecting:
            return
        self.tasks[kw] = {'status': 'pending', 'count': 0, 'data': None, 'error': None}
        self.collecting = True
        self._cancel = False
        self._update_buttons()
        self._render()
        threading.Thread(target=self._run, args=([kw],), daemon=True).start()

    # ── 渲染 ──
    def _render(self):
        t = self.task_table
        t.setRowCount(len(self.keywords))
        for i, kw in enumerate(self.keywords):
            info = self.tasks.get(kw, {})
            st = info.get('status', 'pending')
            cnt = info.get('count', 0)
            cnt_s = f"{cnt:,}" if st in ('done', 'collecting') else '—'

            k0 = QTableWidgetItem(kw)
            k0.setTextAlignment(Qt.AlignCenter)
            t.setItem(i, 0, k0)
            si = QTableWidgetItem(STATUS_L.get(st, st))
            si.setForeground(QColor(STATUS_C.get(st, '#000')))
            si.setTextAlignment(Qt.AlignCenter)
            t.setItem(i, 1, si)
            ci = QTableWidgetItem(cnt_s)
            ci.setTextAlignment(Qt.AlignCenter)
            t.setItem(i, 2, ci)

            # 操作按钮
            ops = QWidget()
            lo = QHBoxLayout(ops)
            lo.setContentsMargins(0, 2, 0, 2)
            lo.setSpacing(2)
            lo.setAlignment(Qt.AlignCenter)

            if st == 'done':
                b1 = self._mk_btn("查看", '#fff', '#1d9bf0', '#1d9bf0', small=True)
                b1.clicked.connect(lambda _, k=kw: self.view_data(k))
                lo.addWidget(b1)
                b2 = self._mk_btn("重采", '#fff', '#536471', '#8899a6', small=True)
                b2.clicked.connect(lambda _, k=kw: self.retry_one(k))
                lo.addWidget(b2)
                b3 = self._mk_btn("导出", '#00ba7c', small=True)
                b3.clicked.connect(lambda _, k=kw: self.export_one(k))
                lo.addWidget(b3)
            elif st == 'failed':
                b1 = self._mk_btn("重采", '#f59e0b', small=True)
                b1.clicked.connect(lambda _, k=kw: self.retry_one(k))
                lo.addWidget(b1)

            if st != 'collecting':
                bd = self._mk_btn("删除", '#fff', '#f4212e', '#f4212e', small=True)
                bd.clicked.connect(lambda _, k=kw: self.remove_kw(k))
                lo.addWidget(bd)

            t.setCellWidget(i, 3, ops)
            t.setRowHeight(i, 40)

    # ── 预览 ──
    def view_data(self, kw):
        self.current_view = kw
        self._show_data(kw)

    def _show_data(self, kw):
        info = self.tasks.get(kw, {})
        data = info.get('data')
        if not data:
            return
        self.preview_box.setVisible(True)
        self.preview_label.setText(f"{kw}（{len(data):,} 条）")

        dt = self.data_table
        dt.clear()
        dt.setColumnCount(len(COLUMNS))
        dt.setHorizontalHeaderLabels(COLUMNS)
        dt.setRowCount(min(len(data), 500))
        for ri, row in enumerate(data[:500]):
            for ci, col in enumerate(COLUMNS):
                # 接口返回什么显示什么（页面为准，不做字典翻译）
                item = QTableWidgetItem(str(row.get(col, '')))
                item.setTextAlignment(Qt.AlignCenter)
                dt.setItem(ri, ci, item)
        # 关键列拉伸，其余固定宽度 + 横向滚动
        for ci, name in enumerate(COLUMNS):
            if name in ('企业名称', '所在市', '统一社会信用代码'):
                dt.horizontalHeader().setSectionResizeMode(ci, QHeaderView.Stretch)
            else:
                dt.setColumnWidth(ci, 120)
        if len(data) > 500:
            dt.setRowCount(501)
            dt.setItem(500, 0, QTableWidgetItem(f'... 仅显示前 500 条，共 {len(data):,} 条'))

    def _hide_preview(self):
        self.preview_box.setVisible(False)
        self.data_table.clear()
        self.data_table.setRowCount(0)

    # ── 导出 ──
    def export_all(self):
        done = [k for k in self.keywords if self.tasks.get(k, {}).get('status') == 'done']
        if not done:
            QMessageBox.warning(self, "提示", "没有可导出的数据")
            return

        box = QMessageBox(self)
        box.setWindowTitle("导出方式")
        box.setText("选择导出方式：")
        merge_btn = box.addButton("合并到一个文件", QMessageBox.YesRole)
        sep_btn = box.addButton("导出独立文件", QMessageBox.NoRole)
        cancel_btn = box.addButton("取消", QMessageBox.RejectRole)
        box.exec_()
        clicked = box.clickedButton()
        if clicked == cancel_btn:
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'export_{ts}')
        os.makedirs(default_dir, exist_ok=True)

        if clicked == merge_btn:
            default_path = os.path.join(default_dir, f'合并_{ts}.xlsx')
            path, _ = QFileDialog.getSaveFileName(self, "保存", default_path, "Excel (*.xlsx)")
            if not path or not path.strip():
                return
            # 确保路径是绝对路径
            if not os.path.isabs(path):
                path = os.path.abspath(path)
            try:
                all_rows, dup = dedup_rows(self.tasks, done)
                if not all_rows:
                    QMessageBox.warning(self, "提示", "去重后没有数据可导出")
                    return
                result_path = write_xlsx_file(path, all_rows)
                # 验证文件是否真的被创建
                if not os.path.exists(result_path):
                    raise RuntimeError(f"文件写入失败，文件不存在: {result_path}")
                file_size = os.path.getsize(result_path)
                if file_size == 0:
                    raise RuntimeError(f"文件大小为0，写入可能失败: {result_path}")
                raw = sum(len(self.tasks[k]['data']) for k in done)
                QMessageBox.information(self, "提示", f"原始 {raw} 条，去重后 {len(all_rows)} 条（去除 {dup} 条重复）\n文件大小: {file_size:,} 字节\n已导出到:\n{path}")
            except RuntimeError as e:
                QMessageBox.critical(self, "导出失败", str(e))
            except Exception as e:
                QMessageBox.critical(self, "导出失败", f"未知错误: {e}")
        else:
            folder = QFileDialog.getExistingDirectory(self, "选择导出文件夹", default_dir)
            if not folder:
                return
            count = 0
            errors = []
            for k in done:
                safe_name = re.sub(r'[\\/:*?"<>|\r\n]', '_', k)[:50]
                fpath = os.path.join(folder, f'{safe_name}_{ts}.xlsx')
                try:
                    write_xlsx_file(fpath, self.tasks[k]['data'])
                    count += 1
                except RuntimeError as e:
                    errors.append(f"{k}: {e}")
                except Exception as e:
                    errors.append(f"{k}: {e}")
            if errors:
                QMessageBox.warning(self, "导出完成（部分失败）", f"成功 {count} 个\n失败 {len(errors)} 个:\n" + "\n".join(errors[:5]))
            else:
                QMessageBox.information(self, "提示", f"已导出 {count} 个文件到:\n{folder}")

    def export_one(self, kw):
        info = self.tasks.get(kw, {})
        data = info.get('data')
        if not data:
            QMessageBox.warning(self, "提示", "该关键词没有可导出的数据")
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'export_{ts}')
        os.makedirs(default_dir, exist_ok=True)
        safe_name = re.sub(r'[\\/:*?"<>|\r\n]', '_', kw)[:50]
        default_path = os.path.join(default_dir, f'{safe_name}_{ts}.xlsx')
        path, _ = QFileDialog.getSaveFileName(self, "导出", default_path, "Excel (*.xlsx)")
        if not path or not path.strip():
            return
        # 确保路径是绝对路径
        if not os.path.isabs(path):
            path = os.path.abspath(path)
        try:
            result_path = write_xlsx_file(path, data)
            # 验证文件是否真的被创建
            if not os.path.exists(result_path):
                raise RuntimeError(f"文件写入失败，文件不存在: {result_path}")
            file_size = os.path.getsize(result_path)
            if file_size == 0:
                raise RuntimeError(f"文件大小为0，写入可能失败: {result_path}")
            QMessageBox.information(self, "提示", f"导出成功\n文件大小: {file_size:,} 字节\n文件: {path}")
        except RuntimeError as e:
            QMessageBox.critical(self, "导出失败", str(e))
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"未知错误: {e}")

    def export_single(self):
        if self.current_view:
            self.export_one(self.current_view)

    # ── 按钮状态 ──
    def _update_buttons(self):
        if self.collecting:
            self.btn_start.setText("采集中...")
            self.btn_add.setEnabled(False)
            self.btn_clear.setEnabled(False)
            self.kw_input.setEnabled(False)
            self.btn_export.setEnabled(False)
        else:
            has_done = any(self.tasks.get(k, {}).get('status') == 'done' for k in self.keywords)
            self.btn_start.setText("▶ 采集")
            self.btn_add.setEnabled(True)
            self.btn_clear.setEnabled(True)
            self.kw_input.setEnabled(True)
            self.btn_export.setEnabled(has_done)


# ======================= 入口 =======================
def main():
    args = [a for a in sys.argv[1:]]
    fresh = "--fresh" in args
    keywords = [a for a in args if not a.startswith("--")]

    # CLI 模式：有参数 → 直接采集导出（支持断点续传）
    if keywords:
        try:
            run_cli(keywords, fresh=fresh)
        except KeyboardInterrupt:
            print("\n⚠ 已中断。已完成关键词的结果已节点保存（采集缓存/），")
            print("  重新运行原命令即可续传，跳过已完成关键词。")
        return

    # 界面模式
    if not HAS_GUI:
        print("界面模式需要 PyQt5，请先安装：pip install PyQt5")
        print("CLI 模式用法：python main.py 关键词1 关键词2 ...")
        print("    --fresh 强制全量重采（默认断点续传）")
        return

    app = QApplication(sys.argv)
    f_ = QFont("Microsoft YaHei"); f_.setPixelSize(14)
    app.setFont(f_)
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
