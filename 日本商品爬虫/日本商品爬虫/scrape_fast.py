#!/usr/bin/env python3
"""
乐天书店 リーメント 商品快速抓取脚本
只抓取列表页的基本信息
"""

import os
import re
import time
import json
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# 尝试导入openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    import subprocess
    subprocess.run(['pip3', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill


class RakutenScraper:
    def __init__(self, base_dir):
        self.base_dir = base_dir
        self.data_file = os.path.join(base_dir, '商品数据_基础.xlsx')
        self.json_file = os.path.join(base_dir, '商品数据_基础.json')
        
        # 请求头
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'ja,en-US;q=0.7,en;q=0.3',
        }
        
        # 数据存储
        self.products = []
        self.session = requests.Session()
    
    def get_page(self, url, retries=3):
        """获取页面内容"""
        for i in range(retries):
            try:
                resp = self.session.get(url, headers=self.headers, timeout=30)
                resp.raise_for_status()
                return resp.text
            except Exception as e:
                if i < retries - 1:
                    time.sleep(1)
        return None
    
    def parse_search_page(self, html):
        """解析搜索结果页面"""
        soup = BeautifulSoup(html, 'html.parser')
        items = soup.find_all('div', class_='rbcomp__item-list__item')
        
        products = []
        for item in items:
            try:
                product = {}
                
                # 商品链接
                link = item.find('a', href=lambda x: x and '/rb/' in x)
                if link:
                    product['url'] = link['href'].split('?')[0]
                    product['id'] = product['url'].split('/rb/')[-1].rstrip('/')
                
                # 商品标题
                title = item.find('span', class_='rbcomp__item-list__item__title')
                product['title'] = title.text.strip() if title else ''
                
                # 分类
                category = item.find('span', class_='rbcomp__category')
                product['category'] = category.text.strip() if category else ''
                
                # JAN码
                isbn = item.find('p', class_='rbcomp__item-list__item__isbn')
                if isbn:
                    jan_text = isbn.text.strip()
                    jan_match = re.search(r'JAN[：:]\s*(\d+)', jan_text)
                    product['jan'] = jan_match.group(1) if jan_match else ''
                
                # 发售日和品牌
                subtexts = item.find_all('p', class_='rbcomp__item-list__item__subtext')
                for st in subtexts:
                    text = st.text.strip()
                    if '発売' in text or 'リーメント' in text:
                        product['release_info'] = text
                        # 提取发售日
                        date_match = re.search(r'(\d{4}年\d{1,2}月\d{1,2}日)', text)
                        product['release_date'] = date_match.group(1) if date_match else ''
                        # 提取品牌
                        if 'リーメント' in text:
                            product['brand'] = 'リーメント'
                        break
                
                # 原价
                original_price = item.find('span', class_='rbcomp__line-through')
                product['original_price'] = original_price.text.strip().replace('円', '') if original_price else ''
                
                # 折扣
                discount = item.find('span', class_='rbcomp__price')
                if discount:
                    em = discount.find('em')
                    product['discount'] = em.text.strip() if em else ''
                
                # 现价
                price = item.find('span', class_='rbcomp__item-list__item__price')
                if price:
                    em = price.find('em')
                    product['price'] = em.text.strip().replace('円', '') if em else ''
                
                # 库存状态
                stock = item.find('p', class_='rbcomp__item-list__item__stock')
                if stock:
                    em = stock.find('em')
                    product['stock_status'] = em.text.strip() if em else ''
                
                # 图片URL
                img = item.find('img')
                if img:
                    product['thumbnail_url'] = img.get('src', '')
                
                if product.get('url'):
                    products.append(product)
                    
            except Exception as e:
                continue
        
        return products
    
    def scrape_all(self, max_pages=None):
        """抓取所有商品"""
        base_url = 'https://books.rakuten.co.jp/search'
        
        page = 1
        total_products = 0
        
        print("=" * 60)
        print("开始抓取リーメント商品数据（快速模式）")
        print("=" * 60)
        
        while True:
            if max_pages and page > max_pages:
                break
            
            print(f"\n正在抓取第 {page} 页...")
            
            # 构建URL
            url = f"{base_url}?seller=%E3%83%AA%E3%83%BC%E3%83%A1%E3%83%B3%E3%83%88&h=100&l-id=search-c-number-03"
            if page > 1:
                url += f"&p={page}"
            
            html = self.get_page(url)
            if not html:
                print("  获取页面失败，停止抓取")
                break
            
            # 解析商品列表
            products = self.parse_search_page(html)
            if not products:
                print("  没有找到更多商品，停止抓取")
                break
            
            print(f"  找到 {len(products)} 个商品")
            
            # 添加到总列表
            self.products.extend(products)
            total_products += len(products)
            
            print(f"  已抓取 {total_products} 个商品")
            
            # 保存进度
            self.save_json()
            
            # 检查是否有下一页
            soup = BeautifulSoup(html, 'html.parser')
            count_text = soup.find(string=lambda x: x and '件' in str(x) and '全' in str(x) if x else False)
            if count_text:
                match = re.search(r'全\s*(\d+)\s*件', count_text)
                if match:
                    total_count = int(match.group(1))
                    print(f"  总计: {total_count} 件")
                    if total_products >= total_count:
                        print("\n已抓取所有商品")
                        break
            
            page += 1
            time.sleep(0.5)  # 页面间延迟
        
        print(f"\n抓取完成！共抓取 {total_products} 个商品")
        self.save_excel()
        return self.products
    
    def save_json(self):
        """保存为JSON（紧凑格式）"""
        with open(self.json_file, 'w', encoding='utf-8') as f:
            json.dump(self.products, f, ensure_ascii=False, separators=(',', ':'))
    
    def save_excel(self):
        """保存为Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'リーメント商品数据'
        
        # 表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # 表头
        headers = [
            '商品ID', '商品名称', '分类', 'JAN码', '发售日', '品牌',
            '原价', '折扣', '现价', '库存状态', '商品链接', '缩略图URL'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 数据
        for row, product in enumerate(self.products, 2):
            ws.cell(row=row, column=1, value=product.get('id', ''))
            ws.cell(row=row, column=2, value=product.get('title', ''))
            ws.cell(row=row, column=3, value=product.get('category', ''))
            ws.cell(row=row, column=4, value=product.get('jan', ''))
            ws.cell(row=row, column=5, value=product.get('release_date', ''))
            ws.cell(row=row, column=6, value=product.get('brand', ''))
            ws.cell(row=row, column=7, value=product.get('original_price', ''))
            ws.cell(row=row, column=8, value=product.get('discount', ''))
            ws.cell(row=row, column=9, value=product.get('price', ''))
            ws.cell(row=row, column=10, value=product.get('stock_status', ''))
            ws.cell(row=row, column=11, value=product.get('url', ''))
            ws.cell(row=row, column=12, value=product.get('thumbnail_url', ''))
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 40
        ws.column_dimensions['L'].width = 40
        
        wb.save(self.data_file)
        print(f"\n数据已保存到: {self.data_file}")


def main():
    base_dir = '/Users/hong/Desktop/日本商品'
    scraper = RakutenScraper(base_dir)
    
    # 抓取所有商品
    products = scraper.scrape_all()
    
    print(f"\n抓取完成！共 {len(products)} 个商品")
    print(f"Excel文件: {scraper.data_file}")
    print(f"JSON文件: {scraper.json_file}")


if __name__ == '__main__':
    main()
